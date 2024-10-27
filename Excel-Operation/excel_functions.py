import pandas as pd
from openpyxl import load_workbook


# read specific column(exclude column name)
def read(file_path, sheet_name, column_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    column_data = df[column_name]
    return column_data


# start from 2 row,insert to last saved page,column start 1
def write_by_index(file_path, column_index, data_list):
    wb = load_workbook(file_path)
    ws = wb.active
    start_row = 1
    for string in data_list:
        start_row = start_row + 1
        cell = ws.cell(row=start_row, column=column_index)
        cell.value = string
    wb.save(file_path)


def write(file_path, sheet_name, column_name, new_data):
    # 加载Excel工作簿
    workbook = load_workbook(file_path)

    # 检查工作表是否存在
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表 '{sheet_name}' 不存在于文件 '{file_path}' 中。")

    # 选择指定工作表
    sheet = workbook[sheet_name]

    # 获取表头和列索引
    header = [cell.value for cell in sheet[1]]  # 假设第一行为表头
    if header and header[0] == column_name:
        col_index = 1  # 列'A'的索引
    else:
        col_index = None
        for i, name in enumerate(header):
            if name == column_name:
                col_index = i + 1  # 1-based index
                break

    if col_index is None:
        raise ValueError(f"列名 '{column_name}' 不在工作表 '{sheet_name}' 中。")

    # 从第二行开始写入数据
    for i, value in enumerate(new_data):
        sheet.cell(row=i + 2, column=col_index, value=value)  # 从第二行开始写入

    # 保存修改
    workbook.save(file_path)
    print(f"数据成功写入 {file_path} 的 '{sheet_name}' 工作表中的 '{column_name}' 列。")


def append(file_path, sheet_name, column_name, new_data):
    # 加载Excel工作簿
    workbook = load_workbook(file_path)

    # 检查工作表是否存在
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表 '{sheet_name}' 不存在于文件 '{file_path}' 中。")

    # 选择指定工作表
    sheet = workbook[sheet_name]

    # 获取表头和列索引
    header = [cell.value for cell in sheet[1]]  # 假设第一行为表头
    if column_name not in header:
        raise ValueError(f"列名 '{column_name}' 不在工作表 '{sheet_name}' 中。")

    # 获取列的索引 (1-based)
    col_index = header.index(column_name) + 1

    # 查找该列最后一个非空单元格的行数
    last_row = sheet.max_row + 1  # 默认下一行是max_row + 1

    # 添加新数据到指定列
    for i, value in enumerate(new_data):
        # 填写新的数据到最后一行之后
        sheet.cell(row=last_row + i, column=col_index, value=value)

    # 保存修改
    workbook.save(file_path)
    print(f"数据成功添加到 {file_path} 的 '{sheet_name}' 工作表中的 '{column_name}' 列。")
