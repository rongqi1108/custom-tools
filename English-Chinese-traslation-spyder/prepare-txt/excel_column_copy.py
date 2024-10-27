import pandas as pd

def read_excel_column(file_path, sheet_name, column_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    column_data = df[column_name]
    return column_data
from openpyxl import load_workbook

def append_strings_to_column(file_path, column_index, strings):
    wb = load_workbook(file_path)
    ws = wb.active
    start_row = 1
    for string in strings:
        start_row = start_row + 1
        cell = ws.cell(row=start_row, column=column_index)
        cell.value = string

    wb.save(file_path)


if __name__ == "__main__":
    file_path = 'C:\\Users\\admin\\Desktop\\vocabularies.xlsx'
    sheet_name = 'Sheet1'
    column_name = '释义' # read column
    column_data = read_excel_column(file_path, sheet_name, column_name)
    column_index = 6  # write column index
    append_strings_to_column(file_path, column_index, column_data)
