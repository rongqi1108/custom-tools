from openpyxl import load_workbook

def append_strings_to_column(file_path, column_index, strings):
    wb = load_workbook(file_path)
    ws = wb.active
    start_row = 1
    for string in strings:
        start_row = start_row + 1
        cell = ws.cell(row=start_row, column=column_index)
        cell.value = string

    wb.save(file_path)

# 示例用法
file_path = 'your_excel_file.xlsx'
column_index = 3  # 假设要追加到第一列
strings = ['str1', 'ng2','string3']
append_strings_to_column(file_path, column_index, strings)