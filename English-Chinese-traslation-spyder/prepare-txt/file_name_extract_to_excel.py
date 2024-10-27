import os
import re
from openpyxl import load_workbook


def append_strings_to_column(file_path, column_index, strings, start_row):
    wb = load_workbook(file_path)
    ws = wb.active
    start_row -= 1
    for string in strings:
        start_row = start_row + 1
        cell = ws.cell(row=start_row, column=column_index)
        cell.value = str(string)
    wb.save(file_path)

def get_all_file_name_from_folder(folder_path):
    file_names = os.listdir(folder_path)
    file_path_list = []
    for file_name in file_names:
        file_path_list.append(file_name)
    return file_path_list

int
if __name__ == '__main__':
    folder_path = 'F:\\英语学习\\S2A4 多情景记单词 雅思托福核心词'
    file_path = 'C:\\Users\\admin\\Desktop\\生词本.xlsx'
    file_name_list = get_all_file_name_from_folder(folder_path)
    size = 1000
    my_list = [''] * size
    for file_name in file_name_list:
        index = int(re.search(r'\d+', file_name).group())
        word = re.search(r'\.(.*?)\.', file_name)
        my_list[index] = word.group(1)

    print(len(my_list))
    append_strings_to_column(file_path, 1, my_list, 2)
