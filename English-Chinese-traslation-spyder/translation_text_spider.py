import math

import requests
import os
from lxml import etree
import re
import socket
import pandas as pd

# 设置请求超时时间，防止长时间停留在同一个请求
socket.setdefaulttimeout(10)

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


class Vocabulary:
    def __init__(self, en, us, uk, simply_ch, ch, ph, se):
        self.en = en
        self.us = us
        self.uk = uk
        self.ch = ch
        self.ph = ph
        self.se = se
        self.simply_ch = simply_ch


def truncate_string(s):
    if '；' in s:
        return s.split('；')[0]
    else:
        return s


def query_from_youdao(keyword):
    header = {
        "Cookie": "OUTFOX_SEARCH_USER_ID=1959936360@10.110.96.158; OUTFOX_SEARCH_USER_ID_NCOO=1296622016.6470625; __yadk_uid=x7bI3vhAaYCied973blqBic7l0nJuOSH; rollNum=true; ___rl__test__cookies=1678353215491; advertiseCookie=advertiseValue",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.63"
    }
    u = "https://dict.youdao.com/result?word={}&lang=en".format(keyword)
    respond = requests.get(url=u, headers=header)
    reExpTrans = re.compile('英</span><span class="phonetic" data-v-39fab836>/ (.*?) /<')
    reExpTrans2 = re.compile('美</span><span class="phonetic" data-v-39fab836>/ (.*?) /<')
    reExpTrans3 = re.compile('class="trans(-content)?".*?>(.*?)<')
    reExpTrans4 = re.compile('{sentence:"(.*?)"')
    reExpTrans5 = re.compile('sentence-translation":"(.*?)"')
    reExpTrans6 = re.compile('<span class="pos" data-v-8042e1b4>(.*?)<')
    reExpTrans7 = re.compile('class="point" data-v-61ce6cc7>(.*?)<')
    reExpTrans8 = re.compile('class="sen-phrase"[^>]*>(.*?)<')

    cont = respond.content.decode('utf-8')
    phonetic_uk = reExpTrans.findall(cont)
    phonetic_us = reExpTrans2.findall(cont)
    meanings = reExpTrans3.findall(cont)
    sen = reExpTrans4.findall(cont)
    sen_tran = reExpTrans5.findall(cont)
    pos = reExpTrans6.findall(cont)
    phrase = reExpTrans7.findall(cont)
    phrase_trans = reExpTrans8.findall(cont)

    # 处理含义
    mea = ''
    # 精简版含义
    si_mea = ''
    i = 0
    for p in pos:
        if len(meanings) == i:
            break
        mea = mea + p + meanings[i][1] + '\n'
        si_mea = si_mea + p + truncate_string(meanings[i][1]) + '\n'
        i = i + 1
    if len(meanings) != len(pos) and len(meanings) != i:
        mea = mea + meanings[i][1]

    # 处理短语
    phrase_all = ''
    i = 0
    for ph_one in phrase:
        phrase_all = phrase_all + ph_one + " " + phrase_trans[i] + '\n'
        i = i + 1

    # 处理例句
    sen_all = ''
    i = 0
    sen_tran_len = len(sen_tran)
    for sen_one in sen:
        if i == sen_tran_len:
            break
        sen_all = sen_all + sen_one + sen_tran[i] + '\n'
        i += 1

    return Vocabulary(keyword, phonetic_us, phonetic_uk, si_mea, mea, phrase_all, sen_all)


def read_excel_column(file_path, sheet_name, column_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    column_data = df[column_name]
    return column_data


def append_strings_to_column(file_path, column_index, strings, start_row):
    wb = load_workbook(file_path)
    ws = wb.active
    start_row -= 1
    for string in strings:
        start_row = start_row + 1
        cell = ws.cell(row=start_row, column=column_index)
        cell.value = str(string)
    wb.save(file_path)


def biying_pic_url(num, keyword):
    """
    该函数用于获取必应图片搜索结果的图片链接

    参数：
    num - 要获取的图片数量
    keyword - 搜索的关键词

    返回：
    一个包含图片链接的列表
    """
    pic_url = []  # 用于存储图片链接的列表
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36'}  # 定义请求头
    for i in range((num // 35) + 1):  # 根据指定的图片数量计算需要请求的页数
        page_url = 'https://cn.bing.com/images/async?q={}&first={}&count=35&relo=4&relp=5&cw=1117&ch=689&scenario=ImageBasicHover&datsrc=N_I&layout=ColumnBased&mmasync=1&dgState=c*6_y*1582s1599s1589s1660s1720s1704_i*40_w*172&IG=B3C2B933EAED48A4A82330EC1E7A638B&SFX=2&iid=images.5659'.format(
            keyword, i * 35)  # 构造每页的 URL
        # print(page_url)  # 打印当前请求的 URL
        html = requests.get(page_url, headers=headers).text  # 发送请求获取页面内容
        html = etree.HTML(html)  # 将页面内容转换为 HTML 格式
        conda_list = html.xpath('//a[@class="iusc"]/@m')  # 使用 XPath 提取特定属性的值
        for j in conda_list:  # 遍历提取到的属性值
            img_url = re.search('"murl":"(.*?)"', j).group(1)  # 使用正则表达式提取图片链接
            pic_url.append(img_url)  # 将图片链接添加到列表中
            if len(pic_url) == num:
                break
    return pic_url  # 返回图片链接列表


def down_img(num, keyword, jpg_url):
    """
    该函数用于下载图片

    参数：
    num - 要下载的图片数量
    keyword - 搜索的关键词，用于创建保存图片的文件夹
    """
    pic_url = biying_pic_url(num, keyword)  # 获取图片链接

    if os.path.exists('D:/图片/' + keyword):  # 检查保存图片的文件夹是否存在
        pass
    else:
        os.makedirs('D:/图片/' + keyword)  # 如果不存在则创建

    path = 'D:/图片/'  # 定义图片保存的路径
    for index, i in enumerate(pic_url):  # 遍历图片链接
        try:
            filename = path + keyword + '/' + str(index) + '.jpg'  # 构造图片保存的文件名
            print(filename)  # 打印文件名
            jpg_url.append(os.path.normpath(filename))
            with open(filename, 'wb+') as f:  # 以二进制写入模式打开文件
                f.write(requests.get(i).content)  # 写入图片内容
        except:
            continue  # 如果下载过程中出现异常则跳过


def read_excel_column(file_path, sheet_name, column_name, start_row, num_rows):
    start_row -= 2
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    column_data = df[column_name][start_row: start_row + num_rows]
    return column_data


def build_process(file_path, sheet_name, column_name, batch_num, start_row,
                  all_row_number):
    # 收集各字段的容器
    us_list = []
    uk_list = []
    ch_list = []
    si_ch_list = []
    ph_list = []
    sen_list = []
    if start_row > all_row_number:
        return 0

    words = read_excel_column(file_path, sheet_name, column_name, start_row, batch_num)
    print("正在查询" + str(len(words)) + "个单词……")
    for word in words:
        one = query_from_youdao(word)
        sub = "-"
        us_list.append(one.us if len(one.us) != 0 else sub)
        uk_list.append(one.uk if len(one.uk) != 0 else sub)
        ch_list.append(one.ch if len(one.ch) != 0 else sub)
        si_ch_list.append(one.simply_ch if len(one.simply_ch) != 0 else sub)
        ph_list.append(one.ph if len(one.ph) != 0 else sub)
        sen_list.append(one.se if len(one.se) != 0 else sub)
    append_strings_to_column(file_path, 2, us_list, start_row)
    append_strings_to_column(file_path, 3, uk_list, start_row)
    append_strings_to_column(file_path, 4, si_ch_list, start_row)
    append_strings_to_column(file_path, 5, ch_list, start_row)
    append_strings_to_column(file_path, 6, ph_list, start_row)
    append_strings_to_column(file_path, 7, sen_list, start_row)
    print("成功导入，还剩大约" + str(all_row_number - start_row) + '个')
    if start_row <= all_row_number:
        build_process(file_path, sheet_name, column_name, batch_num, start_row + batch_num,
                      all_row_number)

if __name__ == '__main__':
    # excel地址 具体哪个Sheet 哪一列
    file_path = 'C:\\Users\\admin\\Desktop\\生词本.xlsx'
    sheet_name = 'Sheet1'
    column_name = '单词'
    # 辅助判定当前表已有多少已经处理的数据
    column_name_b = '美音'

    # 每次查询多少个
    batch_num = 200
    # 读取 Excel 文件
    df = pd.read_excel(file_path)
    # 开始行号
    all_row_number = df[df[column_name].notna()][column_name].index.max() + 2
    start_row = df[df[column_name_b].notna()][column_name_b].index.max() + 3
    if math.isnan(start_row):
        start_row = 2
    print('all_number:' + str(all_row_number))
    print( 'start_number:' + str(start_row))
    build_process(file_path, sheet_name, column_name, batch_num, start_row, all_row_number)
    print('---------------全部任务已完成！---------------')


